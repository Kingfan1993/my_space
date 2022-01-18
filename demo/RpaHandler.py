import copy

import xlrd
import os
from openpyxl import load_workbook
import os.path

my_path = os.path.abspath(os.path.dirname(__file__))

excel_path = os.path.join(my_path, '保单号.xlsx')

data_list_path = os.path.join(my_path, 'bx_list.json')

import json

path = "D:\FFOutput\屏幕录像"


def readExcel():
    if not os.path.exists(data_list_path):
        wb = load_workbook(excel_path)
        sheets = wb.worksheets
        sheet1 = sheets[0]
        max_rows = sheet1.max_row

        res = []

        for row in range(2, max_rows + 1):
            var1 = sheet1.cell(row, 1).value
            res.append(var1)
        write_json(res)
        return res
    else:
        return read_json()


def renameFile():
    file_list = os.listdir(path)
    file_list = [i.replace("格式工厂 屏幕录像", "").replace("_", "").replace(".mp4", "") for i in file_list]
    file_list.sort()
    data_list = readExcel()

    if len(file_list) == len(data_list):
        for index, i in enumerate(data_list):
            filename = file_list[index]

            old_name = "格式工厂 屏幕录像" + filename[:8] + "_" + filename[8:] + ".mp4"
            new_name = i[0] + ".mp4"
            os.rename(path + "\\" + old_name, path + "\\" + new_name)
    else:
        print("文件数量不正确")


def write_json(jlist):
    # 将bx列表写入json文件
    with open('bx_list.json', 'w') as f_obj:
        json.dump(jlist, f_obj)


def read_json():
    # 读取存储于json文件中的列表
    with open('bx_list.json', 'r') as f_obj:
        jlist = json.load(f_obj)
        return jlist

if __name__ == '__main__':
    readExcel()
